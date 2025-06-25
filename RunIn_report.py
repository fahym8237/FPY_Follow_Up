import json
from openpyxl import load_workbook

# File paths
json_file_path = "station_report.json"
excel_file_path = "C:\\Users\\F6CAF02\\Desktop\\Automation\\FPY_Follow_Up\\FPY_Follow_Up.xlsx"

# Load JSON data from file
with open(json_file_path, "r") as f:
    station_report = json.load(f)

# Load the Excel workbook
wb = load_workbook(excel_file_path)
ws = wb.active  # Use ws = wb["SheetName"] if you want a specific sheet

# Start from column J (column 10) and row 41
start_col = 10
start_row = 41

# Write headers: Station, Passed, Failed, Error
headers = ["Station","Units", "Passed", "Failed", "Error"]
for col_offset, header in enumerate(headers):
    ws.cell(row=start_row, column=start_col + col_offset, value=header)

# Write each station's data
for row_offset, (station, results) in enumerate(station_report.items(), start=1):
    ws.cell(row=start_row + row_offset, column=start_col + 0, value=station)
    ws.cell(row=start_row + row_offset, column=start_col + 1, value=results.get("Units", ""))
    ws.cell(row=start_row + row_offset, column=start_col + 2, value=results.get("Passed", ""))
    ws.cell(row=start_row + row_offset, column=start_col + 3, value=results.get("Failed", ""))
    ws.cell(row=start_row + row_offset, column=start_col + 4, value=results.get("Error", ""))

# Save the updated workbook
wb.save(excel_file_path)
print("✅ Data from 'station_report.json' has been written to the Excel file.")
