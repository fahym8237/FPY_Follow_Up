import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from config import target_file,nb_week

sheet_name = "FPY"


# Load the workbook and sheet
wb = openpyxl.load_workbook(target_file)
sheet = wb[sheet_name]

# Step 1: Find the column letter for the specified week
row_index = 2
week_column_index = None

for col in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(row=row_index, column=col).value
    if str(cell_value).strip() == nb_week:
        week_column_index = col
        break

if not week_column_index:
    print(f"Week '{nb_week}' not found in row {row_index}.")
    wb.close()
    exit()

column_letter = get_column_letter(week_column_index)
#print(f"The column for week '{nb_week}' is: {column_letter}")

# Step 2: Copy values and styles from F13:F18 to <column_letter>3:<column_letter>8
for i in range(6):  # 6 rows to copy
    source_cell = sheet[f"F{13 + i}"]
    target_cell = sheet[f"{column_letter}{3 + i}"]
    
    # Copy value
    target_cell.value = source_cell.value

    # Copy style safely using copy()
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

#print(f"Copied F13:F18 to {column_letter}3:{column_letter}8")

# Save and close
wb.save(target_file)
wb.close()
