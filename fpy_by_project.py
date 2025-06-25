import openpyxl
from config import unit_column, source_file2, target_file
def update_fpy_report(source_file2, source_sheet, target_file, target_sheet):
    # Load the source workbook and sheet
    src_wb = openpyxl.load_workbook(source_file2, data_only=True)
    src_ws = src_wb[source_sheet]
    
    # Extract data from source file into a dictionary {Station name: (Unit #, FPY pass #, FPY %)}
    fpy_data = {}
    for row in src_ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        station_name, unit_num, fpy_pass, fpy_percent = row
        fpy_data[station_name] = (unit_num, fpy_pass, fpy_percent)
    
    src_wb.close()
    
    # Load the target workbook and sheet
    tgt_wb = openpyxl.load_workbook(target_file)
    tgt_ws = tgt_wb[target_sheet]
    
   
    # Update the target sheet
    for row in range(13, 19): 
        station_name = tgt_ws.cell(row=row, column=3).value  # Get station name from column 2 (B)
        if station_name in fpy_data:
            tgt_ws.cell(row=row, column=unit_column, value=fpy_data[station_name][0])  # Unit # (C)
            tgt_ws.cell(row=row, column=unit_column+1, value=fpy_data[station_name][1])  # FPY pass # (D)
            tgt_ws.cell(row=row, column=unit_column+2, value=fpy_data[station_name][2])  # FPY % (E)
        
    # Save the updated target workbook
    tgt_wb.save(target_file)
    tgt_wb.close()
    
    print("FPY Report updated successfully \u2714")


source_sheet = "Sheet1"  # Change if the actual sheet name is different
target_sheet = "FPY"

update_fpy_report(source_file2, source_sheet, target_file, target_sheet)